from __future__ import annotations

import re
from collections import Counter

from openpyxl.worksheet.worksheet import Worksheet

from parser.merged_cells import get_cell_value
from parser.models import ChangeBlock, ZoneRef
from parser.normalizer import normalize_for_match, normalize_text

DOC_CODE_RE = re.compile(r"\b[А-ЯA-Z]{2,}\.[0-9]{4}\.[0-9]{3,4}\.[0-9]{4,5}\b")
INDEX_CELL_RE = re.compile(r"^\D*(\d{1,3})\D*$")
HEADER_HINTS = (
    "содержание изменения",
    "извещение",
    "указание о заделе",
    "указания о внедрении",
    "дата выпуска",
    "причина",
)
STOP_MARKERS = (
    "применяемость",
    "разослать",
    "составил",
    "проверил",
    "т. контроль",
    "н. контроль",
    "утвердил",
    "предст. заказ.",
    "изменения внес",
    "контрольную копию исправил",
)


def _compact_fragments(values: list[str | None]) -> list[str]:
    compact: list[str] = []
    prev = None
    for value in values:
        text = normalize_text(value)
        if not text:
            continue
        if text == prev:
            continue
        compact.append(text)
        prev = text
    return compact


def _collect_row_cells(sheet: Worksheet, row_idx: int) -> list[str | None]:
    return [get_cell_value(sheet, row_idx, col_idx) for col_idx in range(1, sheet.max_column + 1)]


def _row_text_from_cells(cells: list[str | None]) -> str:
    unique = _compact_fragments(cells)
    return normalize_text(" ".join(unique)) or ""


def _is_header_like(text: str) -> bool:
    lowered = normalize_for_match(text)
    return any(hint in lowered for hint in HEADER_HINTS)


def _is_stop_marker(text: str) -> bool:
    lowered = normalize_for_match(text)
    return any(marker in lowered for marker in STOP_MARKERS)


def _find_table_anchor(sheet: Worksheet, max_scan_rows: int = 200) -> tuple[int | None, int | None, int | None]:
    izm_row = None
    izm_col = None
    content_row = None

    for row_idx in range(1, min(sheet.max_row, max_scan_rows) + 1):
        row_cells = _collect_row_cells(sheet, row_idx)
        for col_idx, value in enumerate(row_cells, start=1):
            norm = normalize_for_match(value)
            if not norm:
                continue
            if izm_row is None and (norm == "изм." or norm == "изм" or norm.startswith("изм.")):
                izm_row, izm_col = row_idx, col_idx
            if content_row is None and "содержание изменения" in norm:
                content_row = row_idx

    if izm_row is None:
        return None, None, None

    # continuation sheets may have compact header where "Содержание изменения" may be omitted/truncated
    if content_row is None:
        content_row = izm_row

    header_end = max(izm_row, content_row)
    return header_end + 1, izm_col, header_end


def _find_doc_code_nearby(sheet: Worksheet, row_idx: int, start_col: int) -> str | None:
    for probe_row in range(row_idx, min(sheet.max_row, row_idx + 2) + 1):
        cells = _collect_row_cells(sheet, probe_row)
        text = _row_text_from_cells(cells[max(0, start_col - 2) :])
        match = DOC_CODE_RE.search(text)
        if match:
            return match.group(0)
    return None


def _extract_meta_signature(sheet: Worksheet, row_idx: int, idx_col: int) -> tuple[str | None, str | None]:
    cells = _collect_row_cells(sheet, row_idx)
    idx_cell_text = normalize_text(cells[idx_col - 1] if idx_col - 1 < len(cells) else None)
    idx_match = INDEX_CELL_RE.match(idx_cell_text or "")
    if not idx_match:
        return None, None
    doc_code = _find_doc_code_nearby(sheet, row_idx, idx_col)
    if not doc_code:
        return None, None
    return idx_match.group(1), doc_code


def _clean_body_line(line: str) -> str | None:
    cleaned = normalize_text(line)
    if not cleaned:
        return None
    cleaned = cleaned.strip()
    if cleaned in {"-", "--", "---"}:
        return None
    return cleaned


def extract_changes(
    sheet: Worksheet,
    sheet_index: int,
    start_global_seq: int,
) -> tuple[list[ChangeBlock], int, dict[str, object]]:
    table_data_start, idx_col, header_row = _find_table_anchor(sheet)
    debug_rejects = Counter()
    potential_rows = 0
    stop_hits = 0
    closed_by_next_meta = 0
    closed_by_stop = 0

    debug_info: dict[str, object] = {
        "table_found": table_data_start is not None and idx_col is not None,
        "table_header_row_start": header_row,
        "table_data_row_start": table_data_start,
        "potential_meta_rows": 0,
        "rejected_meta_rows": 0,
        "reject_reasons": {},
        "stop_markers_hit": 0,
        "blocks_closed_by_next_meta": 0,
        "blocks_closed_by_stop_marker": 0,
    }

    if table_data_start is None or idx_col is None:
        debug_rejects["table_anchor_not_found"] += 1
        debug_info["rejected_meta_rows"] = 1
        debug_info["reject_reasons"] = dict(debug_rejects)
        return [], start_global_seq, debug_info

    starts: list[tuple[int, str, str]] = []
    for row_idx in range(table_data_start, sheet.max_row + 1):
        row_cells = _collect_row_cells(sheet, row_idx)
        row_text = _row_text_from_cells(row_cells)
        if not row_text:
            continue

        idx_cell_text = normalize_text(row_cells[idx_col - 1] if idx_col - 1 < len(row_cells) else None)
        idx_match = INDEX_CELL_RE.match(idx_cell_text or "")
        if not idx_match:
            if "изм" in normalize_for_match(row_text):
                potential_rows += 1
                debug_rejects["invalid_index"] += 1
            continue

        potential_rows += 1
        if _is_header_like(row_text):
            debug_rejects["header_like"] += 1
            continue

        doc_code = _find_doc_code_nearby(sheet, row_idx, idx_col)
        if not doc_code:
            debug_rejects["no_doc_code"] += 1
            continue

        starts.append((row_idx, idx_match.group(1), doc_code))

    changes: list[ChangeBlock] = []
    for seq_on_sheet, (start_row, change_index, doc_code) in enumerate(starts, start=1):
        nominal_end = starts[seq_on_sheet][0] - 1 if seq_on_sheet < len(starts) else sheet.max_row
        real_end = nominal_end

        body_lines: list[str] = []
        prev_line = None
        for row_idx in range(start_row, nominal_end + 1):
            if row_idx > start_row:
                next_idx, next_doc = _extract_meta_signature(sheet, row_idx, idx_col)
                if next_idx and next_doc:
                    real_end = row_idx - 1
                    closed_by_next_meta += 1
                    break

            row_cells = _collect_row_cells(sheet, row_idx)
            line = _row_text_from_cells(row_cells[idx_col:])
            if not line:
                continue
            if _is_stop_marker(line):
                stop_hits += 1
                closed_by_stop += 1
                real_end = row_idx - 1
                break
            if _is_header_like(line):
                continue
            if DOC_CODE_RE.search(line) and row_idx == start_row:
                continue

            cleaned = _clean_body_line(line)
            if not cleaned:
                continue
            if cleaned == prev_line:
                continue
            body_lines.append(cleaned)
            prev_line = cleaned

        changes.append(
            ChangeBlock(
                sheet_index=sheet_index,
                change_seq_global=start_global_seq,
                change_seq_on_sheet=seq_on_sheet,
                change_index=change_index,
                doc_code=doc_code,
                change_text=normalize_text("\n".join(body_lines)),
                raw_meta_text=f"Изм. {change_index} {doc_code}",
                zone_ref=ZoneRef(
                    meta_row_start=start_row,
                    meta_row_end=start_row,
                    body_row_start=start_row,
                    body_row_end=real_end,
                ),
            )
        )
        start_global_seq += 1

    debug_info["potential_meta_rows"] = potential_rows
    debug_info["rejected_meta_rows"] = int(sum(debug_rejects.values()))
    debug_info["reject_reasons"] = dict(debug_rejects)
    debug_info["stop_markers_hit"] = stop_hits
    debug_info["blocks_closed_by_next_meta"] = closed_by_next_meta
    debug_info["blocks_closed_by_stop_marker"] = closed_by_stop
    return changes, start_global_seq, debug_info


__all__ = ["extract_changes", "DOC_CODE_RE"]
