from __future__ import annotations

from functools import lru_cache

from openpyxl.worksheet.worksheet import Worksheet


def _normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


@lru_cache(maxsize=512)
def _merged_anchor_for(sheet_title: str, merged_range: str) -> tuple[int, int, int, int]:
    # cache parser-friendly boundaries: min_col, min_row, max_col, max_row
    start, end = merged_range.split(":")
    # openpyxl uses e.g. A1 style; worksheet indexing is 1-based
    from openpyxl.utils.cell import coordinate_to_tuple, column_index_from_string

    start_col_letters = "".join(ch for ch in start if ch.isalpha())
    end_col_letters = "".join(ch for ch in end if ch.isalpha())
    start_row = coordinate_to_tuple(start)[0]
    end_row = coordinate_to_tuple(end)[0]
    start_col = column_index_from_string(start_col_letters)
    end_col = column_index_from_string(end_col_letters)
    return start_col, start_row, end_col, end_row


def get_cell_value(sheet: Worksheet, row: int, col: int) -> str | None:
    cell = sheet.cell(row=row, column=col)
    raw = _normalize_text(cell.value)
    if raw is not None:
        return raw

    for merged in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = _merged_anchor_for(sheet.title, str(merged))
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return _normalize_text(sheet.cell(row=min_row, column=min_col).value)

    return None


def iter_sheet_rows(sheet: Worksheet, min_col: int = 1, max_col: int | None = None):
    max_col = max_col or sheet.max_column
    for row_idx in range(1, sheet.max_row + 1):
        yield row_idx, [get_cell_value(sheet, row_idx, col_idx) for col_idx in range(min_col, max_col + 1)]
