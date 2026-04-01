import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook

from parser.header_extractor import extract_document_header


def _fill_common_header(ws, base_row: int = 1, shifted: bool = False):
    offset = 1 if shifted else 0
    ws.cell(base_row, 1 + offset, "Предприятие организация")
    ws.cell(base_row, 2 + offset, 'АО "ИЦ КТ"')
    ws.cell(base_row + 1, 1 + offset, "Извещение")
    ws.cell(base_row + 1, 2 + offset, "ИИ.0000.0001")
    ws.cell(base_row + 2, 1 + offset, "Причина")
    ws.cell(base_row + 2, 2 + offset, "Устранение ошибок")
    ws.cell(base_row + 3, 1 + offset, "Код")
    ws.cell(base_row + 3, 2 + offset, "7")
    ws.cell(base_row + 3, 3 + offset, "Лист")
    ws.cell(base_row + 3, 4 + offset, "1")
    ws.cell(base_row + 3, 5 + offset, "Листов")
    ws.cell(base_row + 3, 6 + offset, "1")
    ws.cell(base_row + 4, 1 + offset, "Указание о заделе")
    ws.cell(base_row + 4, 2 + offset, "Задела нет")
    ws.cell(base_row + 5, 1 + offset, "Указания о внедрении")
    ws.cell(base_row + 5, 2 + offset, "-")
    ws.cell(base_row + 6, 1 + offset, "Применяемость")
    ws.cell(base_row + 6, 2 + offset, "-")
    ws.cell(base_row + 7, 1 + offset, "Разослать")
    ws.cell(base_row + 7, 2 + offset, "-")

    ws.cell(base_row + 15, 1 + offset, "Составил")
    ws.cell(base_row + 15, 2 + offset, "Т. контроль")
    ws.cell(base_row + 15, 3 + offset, "Петров")
    ws.cell(base_row + 16, 1 + offset, "Проверил")
    ws.cell(base_row + 16, 2 + offset, "Предст. заказ.")
    ws.cell(base_row + 16, 3 + offset, "Иванов")
    ws.cell(base_row + 17, 1 + offset, "Н. контроль")
    ws.cell(base_row + 17, 2 + offset, "Н. контроль")
    ws.cell(base_row + 17, 3 + offset, "Свердлов")
    ws.cell(base_row + 18, 1 + offset, "Утвердил")
    ws.cell(base_row + 18, 2 + offset, "Утвердил")
    ws.cell(base_row + 18, 3 + offset, "Алербов")


def test_developer_extraction_from_org_cell():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "Предприятие организация")
    ws.cell(1, 2, 'АО "ИЦ КТ"')

    header, _, debug = extract_document_header(ws)
    assert header.developer == 'АО "ИЦ КТ"'
    assert debug["fields"]["developer"]["anchor_cell"] == "R1C1"


def test_code_extraction_from_dedicated_cell_not_sheet_counter():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "Код")
    ws.cell(1, 2, "7")
    ws.cell(1, 3, "Лист")
    ws.cell(1, 4, "1")
    ws.cell(1, 5, "Листов")
    ws.cell(1, 6, "1")

    header, _, debug = extract_document_header(ws)
    assert header.code == "7"
    assert debug["fields"]["code"]["anchor_cell"] == "R1C1"
    assert debug["fields"]["code"]["value_window"][0] == "7"


def test_sheet_no_declared_extraction_from_list_cell():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "Лист")
    ws.cell(1, 2, "1")
    ws.cell(1, 3, "Листов")
    ws.cell(1, 4, "3")

    header, _, debug = extract_document_header(ws)
    assert header.sheet_no_declared == 1
    assert header.sheet_total_declared == 3
    assert debug["fields"]["sheet_no_declared"]["anchor_cell"] == "R1C1"
    assert debug["fields"]["sheet_total_declared"]["anchor_cell"] == "R1C3"


def test_dash_preserving_fields_return_dash():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "Указания о внедрении")
    ws.cell(1, 2, "-")
    ws.cell(2, 1, "Применяемость")
    ws.cell(2, 2, "-")
    ws.cell(3, 1, "Разослать")
    ws.cell(3, 2, "-")

    header, _, debug = extract_document_header(ws)
    assert header.implementation_instruction == "-"
    assert header.applicability == "-"
    assert header.distribution == "-"
    assert "-" in debug["fields"]["implementation_instruction"]["value_window"]


def test_approvals_values_extracted_to_right_of_labels():
    wb = Workbook()
    ws = wb.active
    ws.cell(20, 1, "Составил")
    ws.cell(20, 2, "Составил")
    ws.cell(20, 3, "Петров")
    ws.cell(21, 1, "Проверил")
    ws.cell(21, 2, "Проверил")
    ws.cell(21, 3, "Иванов")
    ws.cell(22, 1, "Н. контроль")
    ws.cell(22, 2, "Н. контроль")
    ws.cell(22, 3, "Свердлов")
    ws.cell(23, 1, "Утвердил")
    ws.cell(23, 2, "Утвердил")
    ws.cell(23, 3, "Алербов")

    _, approvals, debug = extract_document_header(ws)
    assert approvals.author == "Петров"
    assert approvals.reviewer == "Иванов"
    assert approvals.norm_control == "Свердлов"
    assert approvals.approver == "Алербов"
    assert debug["approvals_rejected_reasons"] == {}
    assert debug["approval_block_detected"] is True
    assert debug["approval_anchor_cells"]["author"] == "R20C1"


def test_anchor_based_local_zones_work_for_shifted_layouts():
    wb = Workbook()
    ws1 = wb.active
    _fill_common_header(ws1, base_row=1, shifted=False)
    ws2 = wb.create_sheet("alt")
    _fill_common_header(ws2, base_row=3, shifted=True)

    for ws in (ws1, ws2):
        header, approvals, _ = extract_document_header(ws)
        assert header.developer == 'АО "ИЦ КТ"'
        assert header.notice_number == "ИИ.0000.0001"
        assert header.reason == "Устранение ошибок"
        assert header.code == "7"
        assert header.sheet_no_declared == 1
        assert header.sheet_total_declared == 1
        assert header.stock_instruction == "Задела нет"
        assert header.implementation_instruction == "-"
        assert header.applicability == "-"
        assert header.distribution == "-"
        assert approvals.author == "Петров"
        assert approvals.reviewer == "Иванов"
        assert approvals.norm_control == "Свердлов"
        assert approvals.approver == "Алербов"
