import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook

from parser.sheet_parser import detect_notice_number


def test_notice_number_is_null_when_only_doc_code_present():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "Извещение № ЕСРТ.0016.716.04121")

    value, dbg = detect_notice_number(ws, sheet_kind="full")
    assert value is None
    assert dbg["rejected_notice_candidates"][0]["reason"] == "looks_like_doc_code"


def test_notice_number_extracted_for_non_doc_code_pattern():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "Извещение № 42-17/2026")

    value, _ = detect_notice_number(ws, sheet_kind="full")
    assert value == "42-17/2026"
