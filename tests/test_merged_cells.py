import pytest
openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook

from parser.merged_cells import get_cell_value


def test_get_cell_value_reads_merged_anchor():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "ANCHOR"
    ws.merge_cells("A1:C2")

    assert get_cell_value(ws, 1, 1) == "ANCHOR"
    assert get_cell_value(ws, 2, 3) == "ANCHOR"
