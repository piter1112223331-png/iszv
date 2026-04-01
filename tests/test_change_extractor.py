import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook

from parser.change_extractor import extract_changes


def test_continuation_sheet_table_is_parsed():
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "Лист 2")
    ws.cell(2, 1, "Изм.")
    ws.cell(2, 3, "Содержание изменения")
    ws.cell(4, 1, "1")
    ws.cell(4, 3, "ЕСРТ.0016.716.04121")
    ws.cell(5, 3, "Текст продолжения")

    changes, _, dbg = extract_changes(ws, sheet_index=2, start_global_seq=1, sheet_kind="continuation")

    assert dbg["table_found"] is True
    assert len(changes) == 1
    assert changes[0].change_text == "Текст продолжения"


def test_first_block_under_header_is_not_empty():
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 1, "Изм.")
    ws.cell(2, 3, "Содержание изменения")
    ws.cell(3, 1, "1")
    ws.cell(3, 3, "ЕСРТ.0016.0000.0001 Изменить покрытие")
    ws.cell(5, 1, "2")
    ws.cell(5, 3, "ЕСРТ.0016.0000.0002")
    ws.cell(6, 3, "Второй блок")

    changes, _, dbg = extract_changes(ws, sheet_index=1, start_global_seq=1, sheet_kind="full")

    assert len(changes) >= 1
    assert changes[0].change_text == "Изменить покрытие"
    assert dbg["first_block_detected"] is True
    assert dbg["first_block_body_rows"] >= 1


def test_full_first_block_fallback_recovers_text_from_next_rows():
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 1, "Изм.")
    ws.cell(2, 3, "Содержание изменения")
    ws.cell(3, 1, "2")
    ws.cell(3, 3, "ЕСРТ.0016.0000.0002")
    ws.cell(4, 3, "Разослать")  # closes main body scan early
    ws.cell(5, 3, "Описание первого изменения")
    ws.cell(6, 1, "3")
    ws.cell(6, 3, "ЕСРТ.0016.0000.0003")

    changes, _, dbg = extract_changes(ws, sheet_index=1, start_global_seq=1, sheet_kind="full")

    assert changes[0].change_text == "Описание первого изменения"
    assert dbg["first_block_fallback_used"] is True
    assert dbg["first_block_final_text"] == "Описание первого изменения"


def test_fallback_does_not_capture_next_meta_row():
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 1, "Изм.")
    ws.cell(2, 3, "Содержание изменения")
    ws.cell(3, 1, "2")
    ws.cell(3, 3, "ЕСРТ.0016.0000.0002")
    ws.cell(4, 1, "3")
    ws.cell(4, 3, "ЕСРТ.0016.0000.0003")
    ws.cell(5, 3, "Текст второго изменения")

    changes, _, dbg = extract_changes(ws, sheet_index=1, start_global_seq=1, sheet_kind="full")

    assert changes[0].change_text is None
    assert dbg["first_block_fallback_used"] is False


def test_first_block_with_merged_rows_keeps_body_text():
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 1, "Изм.")
    ws.cell(2, 3, "Содержание изменения")
    ws.cell(3, 1, "1")
    ws.cell(3, 3, "ЕСРТ.0016.716.04121")
    ws.cell(4, 3, "Текст первого изменения")
    ws.merge_cells("C4:E4")

    changes, _, _ = extract_changes(ws, sheet_index=1, start_global_seq=1, sheet_kind="full")

    assert len(changes) == 1
    assert changes[0].change_text == "Текст первого изменения"


def test_stop_markers_cut_body_and_signatures_not_included():
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 1, "Изм.")
    ws.cell(2, 3, "Содержание изменения")
    ws.cell(4, 1, "2")
    ws.cell(4, 3, "ЕСРТ.0016.0000.0001")
    ws.cell(5, 3, "Изменить покрытие")
    ws.cell(6, 3, "Разослать")
    ws.cell(7, 3, "Составил")

    changes, _, dbg = extract_changes(ws, sheet_index=1, start_global_seq=1, sheet_kind="full")

    assert len(changes) == 1
    assert "Разослать" not in (changes[0].change_text or "")
    assert "Составил" not in (changes[0].change_text or "")
    assert dbg["blocks_closed_by_stop_marker"] >= 1


def test_neighbor_meta_row_closes_previous_block():
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 1, "Изм.")
    ws.cell(2, 3, "Содержание изменения")
    ws.cell(4, 1, "1")
    ws.cell(4, 3, "ЕСРТ.0016.716.04121")
    ws.cell(5, 3, "Первый блок")
    ws.cell(6, 1, "2")
    ws.cell(6, 3, "ЕСРТ.0016.716.04122")
    ws.cell(7, 3, "Второй блок")

    changes, _, dbg = extract_changes(ws, sheet_index=1, start_global_seq=1, sheet_kind="full")

    assert len(changes) == 2
    assert changes[0].change_text == "Первый блок"
    assert changes[1].change_text == "Второй блок"


def test_trailing_dash_cleanup():
    wb = Workbook()
    ws = wb.active
    ws.cell(2, 1, "Изм.")
    ws.cell(2, 3, "Содержание изменения")
    ws.cell(4, 1, "1")
    ws.cell(4, 3, "ЕСРТ.0016.716.04121")
    ws.cell(5, 3, "Изменить материал")
    ws.cell(6, 3, "-")

    changes, _, _ = extract_changes(ws, sheet_index=1, start_global_seq=1, sheet_kind="full")

    assert changes[0].change_text == "Изменить материал"
